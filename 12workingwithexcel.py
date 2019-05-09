#--------------------------------------------------------#
#               Ch. 12 Working with Excel                #
#--------------------------------------------------------#

import os
os.chdir("C:\\Users\\ejmil\\source\\repos\\AutoTheBoringStuff")

import openpyxl as op

wb = op.load_workbook('example.xlsx')
type(wb)

#openpyxl.load_workbook instiates a Workbook object

### Getting Sheets from the Workbook ###
import openpyxl as op
wb = op.load_workbook('example.xlsx')
wb.get_sheet_names()                    #creates a list of workbooks 

sheet = wb.get_sheet_by_name('Sheet3')  #creates a Worksheet object of called worksheet
sheet

type(sheet)

anotherSheet = wb.active                #creates a Worksheet object of the active worksheet

### Getting Cells from the Sheets ###
import openpyxl as op
from openpyxl.utils import get_column_letter

wb = op.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

sheet['A1']                             #creates a Cell object of the cell A1

sheet['A1'].value                       #shows values in that cell

c = sheet['B1']
c.value
c.row
c.column

'Row ' + str(c.row) + ', Column ' + get_column_letter(c.column) + ' is ' + c.value 
#.row gets row num // .column gets column num (use get_colum_letter) from openpyxl.util to get the letter

'Cell ' + c.coordinate + ' is ' + c.value               #.coordinate gives coordinate of value

sheet['C1'].value                                       #.value gives values within cell

#by default columns will be by number because after Z there will be double Letters which gets confusing

sheet.cell(row=1, column=2)                     #cell() method and passing row/column attr will work like sheet['cell']  
sheet.cell(row=1, column=2).value
for i in range(1, 4):
    print(i, sheet.cell(row=2, column=i).value)

import openpyxl as op
wb = op.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet.max_row                               #.max_row gets maximum number of rows in called sheet
sheet.max_column                            #.max_column gets maxiumum number of rows in called sheet (int rather than a letter)

### Converting between Column Letters and Numbers ###

import openpyxl

get_column_letter(1)                        #get_column_letter gets the letter(s) of a column from an entered number
get_column_letter(27)
get_column_letter(2)
get_column_letter(900)

import openpyxl
from openpyxl.utils import column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

get_column_letter(sheet.max_column)
column_index_from_string('A')               #from open.utils.column_index_from_string(string) gets column number from an entered string
column_index_from_string('AA')

### Gettting Rows and Columns from the Sheets ###

#Can slice rows or columns from a Worksheet
#Can also select a rectangular area of a Worksheet

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:               #loops for each row
    for cellObj in rowOfCellObjects:                    #loops for each cell
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')

#To access a particular row or column use the Worksheet object's row and columns attributes
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

sheet['B']
for cellObj in sheet['B']:
    print(cellObj.value)

sheet[1]
for cellObj in sheet[1]:
    print(cellObj.value)

### Creating and Saving Excel Documents ###
import openpyxl
wb = openpyxl.Workbook()                #creates a new workbook
wb.get_sheet_names()
sheet = wb.active
sheet.title                             #.title displays title of sheet
sheet.title = 'Spam Bacon Eggs Sheet'   #.title = changes title
wb.get_sheet_names()

#changes will only be saved if save() is called
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')

### Creating and Removing Sheets ###
#create_sheet() and remove_sheet() methods

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
wb.create_sheet()
wb.sheetnames
wb.create_sheet(index = 0, title = 'First Sheet')           #index= is for placement of sheet; title= is used to name the sheet
wb.create_sheet(index = 2, title = 'Middle Sheet')
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
wb.sheetnames

#remove_sheet() takes a Worksheet object NOT a string. 
    #easiest way is to call .sheetname('name') within remove_sheet()

### Writing Values to Cells ### 
#like writing values to keys in a dictionary

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hellow world!'
sheet['A1'].value

### Setting the Font Style of Cells ###
#FONT() function from the openpyxl.styles module can help stylize spreadsheets
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
italic25Font = Font(size= 24, italic= True)
sheet['A1'].font = italic25Font
sheet['A1'] = 'Hello World!'
wb.save('stylish.xlsx')

### Font Objects ###

#Key Arg#           #DataType#          #Description#
#name               string              The font name, such as 'Calibri' or 'Times New Roman'
#size               integer             The point size
#bold               boolean             =True, for bold font
#italic             boolean             =True, for italic font

#Call Font() to create a Font object as a variable
#Pass that to Style(), store the Style object as a variable and assign this to a Cell objects style attr

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

fontObj1 = Font(name= 'Times New Roman', bold= True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Testing Testing'

fontObj2 = Font(size= 24, italic= True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')

### Formulas ###
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'                 # enter a formula just like text into a cell
wb.save('writeFormula.xlsx')

### Adjusting Rows and Columns ###
#Worksheet objects have row_dimensions and column_dimensions attributes

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall Row'
sheet['B2'] = 'Wide Column'
sheet.row_dimensions[1].height = 70          #.row_dimensions are like dictionary 
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

### Freeze Panes ###
#useful for spreadsheets that are too large. You can freeze headers so you know what the rows and columns are
#Worksheet object has a freez_panes attr that can be applied to a cell
    #all rows above and cells to the left of this cell will be froze. 

#freeze_pane settings#                  #rows and columns frozen#
#sheet.freeze_panes = 'A1'              Row 1
#sheet.freeze_panes = 'B1'              Column A
#sheet.freeze_panes = 'C1'              Columns A and B
#sheet.freeze_panes = 'C2'              Row 1 and Columns A and B
#sheet.freeze_panes = '1' or None       No panes will be frozen

import openpyxl
wb= openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezePanesExample.xlsx')

### Charts ###
#Supports creating bar, line, scatter and pie charts using data in excel

#To create a chart you must..
    #1. Create a reference object from a rectangular selection of cells
    #2. Create a Series object by passing in the Reference object.
    #3. Create a Chart object
    #4. Append the series object to the Chart object
    #5. Add the Chart object to the Worksheet object, optionally specifying which cell the top left corner of the chart should be positioned.

#Reference objects are created by calling openpyxl.chart.Reference() function
    #is passed 3 arguments
        #1. The worksheet object containing your chart
        #2. A tuple of two integers, representing the top left cell of the rectangular selection of cells containing your
            #chart data: the first int in the tup is the row, the second is the column.
        #3. A tuple of two integers, representing the bottom-right cell. of the rectangular selection of cells contain your chart data:
            # The first int in the tup is the row and second is the column

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col= 1, min_row= 1, max_col= 1, max_row= 10)

seriesObj= openpyxl.chart.Series(refObj, title= 'First Series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')