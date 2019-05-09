###-----------------------------------###
#      Ch 12 Updating a Spreadsheet     # 
###-----------------------------------###
"""

This program must...
Loop over all rows and if the row is for garlic, celery or lemons then change the price.

"""
#! python3
import openpyxl

# Open the work book
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# Create dictionary for updated prices
UPDATED_PRICES = {'Celery': 1.19,
                 'Garlic': 3.07,
                 'Lemon': 1.27}

#TODO: Loop through each row, check for Garlic, Lemon and Celery, and update prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in UPDATED_PRICES:
        sheet.cell(row = rowNum, column = 2).value = UPDATED_PRICES[produceName]

#TODO: Save to new workbook
wb.save('updatedPrices.xlsx')
